import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

class ExcelManager:
    """
    This class handles the creation and management of an Excel file for tracking payments.
    """

    def __init__(self, file_path=None):
        """
        Constructor for the ExcelManager class.
        Ensures the correct file path is used and creates the Excel file if missing.
        """
        base_dir = os.path.expanduser("~/Documents/PRA_Records")  # Save records in the user's Documents folder

        if not os.path.exists(base_dir):
            os.makedirs(base_dir)  # Create the directory if it does not exist

        self.file_path = os.path.join(base_dir, "payment_records.xlsx")  # Set the file path
        print(f"✅ EXCEL FILE IS SAVED AT: {self.file_path}")

        # 🛠 ** New added control.**: If the file does not exist, creat it automatically
        if not os.path.exists(self.file_path):
            print("⚠️ Excel file not found, creating a new one...")
            self.create_excel_file()
    
    def create_excel_file(self):
        """
        Creates a new Excel file and initializes the headers.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Records"

        # Define headers for the Excel file
        headers = [
            "Invoice No", "Task Type", "Tariff Fee", "Gross Fee (TL)", "VAT (%)",
            "VAT Amount (TL)", "Net Fee (TL)", "Case Details", "Submission Date",
            "Invoice Date", "Payment Status"
        ]
        ws.append(headers)

        # Save the Excel file
        wb.save(self.file_path)
        print(f"✅ New Excel file created: {self.file_path}")

    def load_workbook(self):
        """
        Loads the existing Excel file and returns the workbook object.
        """
        return load_workbook(self.file_path)

    def add_payment(self, payment_data):
        """
        Adds a new payment record to the Excel file.
        """
        wb = self.load_workbook()
        ws = wb.active
        ws.append(payment_data)  # Append new row
        wb.save(self.file_path)
        wb.close()
        print("✅ New payment record added.")

    def adjust_excel_formatting(self):
        """
        Adjusts column widths and row heights dynamically.
        Ensures proper text wrapping and applies TL currency formatting.
        """
        wb = self.load_workbook()
        ws = wb.active

        # Task Type column letter (B sütunu)
        task_type_column_letter = "B"

        # Adjust column widths based on content
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter (A, B, C...)

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Limit the column width to 30 characters, force wrapping
            if col_letter == task_type_column_letter:
                ws.column_dimensions[col_letter].width = 30  # Fixed width
            else:
                ws.column_dimensions[col_letter].width = max_length + 2  # Apply precise width

        # Apply TL format and adjust row heights dynamically
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            max_height = 15  # Default row height
            for idx, cell in enumerate(row):
                if idx in [2, 3, 5, 6]:  # Columns: Tariff Fee, Gross Fee, VAT Amount, Net Fee
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 TL'  # Set currency format
                        cell.value = float(cell.value)  # Ensure numeric format

                # Force text wrapping
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

                # Only for Task Type column (B)
                if cell.column_letter == task_type_column_letter:
                    text_length = len(str(cell.value)) if cell.value else 0
                    lines = (text_length // 30) + 1  # Wrap after 30 characters
                    max_height = max(max_height, lines * 15)  # Adjust row height

            ws.row_dimensions[row_idx].height = max_height  # Apply final row height

        wb.save(self.file_path)
        wb.close()
        print("✅ Excel formatting adjusted: column widths, row heights, and TL format applied!")

    def update_payment_status(self, invoice_no, new_status):
        """
        Updates the payment status (Pending <-> Paid) in the Excel file.
        """
        wb = self.load_workbook()
        ws = wb.active
        found = False

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == invoice_no:
                row[-1].value = new_status  # Update status in the last column
                found = True
                break

        if found:
            wb.save(self.file_path)
            wb.close()
            return True
        else:
            wb.close()
            return False


    def search_payment(self, invoice_no):
        """
        Searches for a payment by Invoice No in the Excel file.
        Returns payment details if found, otherwise returns None.
        """
        wb = self.load_workbook()
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] == invoice_no:
                wb.close()
                return row  # Return the payment record

        wb.close()
        return None  # Return None if not found

    def list_payments(self):
        """
        Displays all recorded payments in the console.
        """
        wb = self.load_workbook()
        ws = wb.active

        print("\n📌 Recorded Payments:")
        for row in ws.iter_rows(values_only=True):
            print(row)
    
    def analyze_payments(self):
        """
        Analyzes payment records and provides statistical insights.
        Returns key statistics to be displayed in GUI.
        """
        wb = self.load_workbook()
        ws = wb.active

        total_payments = 0
        total_paid = 0
        total_pending = 0
        sum_paid_amounts = 0
        sum_pending_amounts = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[-1] == "Paid":
                total_paid += 1
                sum_paid_amounts += row[6]  # Net Fee column (index 6)
            elif row[-1] == "Pending":
                total_pending += 1
                sum_pending_amounts += row[6]  # Net Fee column (index 6)

            total_payments += 1

        avg_paid = sum_paid_amounts / total_paid if total_paid > 0 else 0
        avg_pending = sum_pending_amounts / total_pending if total_pending > 0 else 0

        wb.close()
        return total_payments, total_paid, sum_paid_amounts, avg_paid, total_pending, sum_pending_amounts, avg_pending

    def highlight_payments(self):
        """
        Highlights 'Paid' payments in green and 'Pending' payments in red in the Excel file.
        """
        wb = self.load_workbook()
        ws = wb.active

        # Define color fills
        green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Light red

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_cell = row[-1]  # Payment Status column (last column)

            if status_cell.value == "Paid":
                status_cell.fill = green_fill  # Apply green fill
            elif status_cell.value == "Pending":
                status_cell.fill = red_fill  # Apply red fill

        wb.save(self.file_path)
        wb.close()
        print("✅ Payment statuses highlighted in Excel!")      

    def get_all_payments(self):
        """
        Retrieves all payment records from the Excel file.
        Returns a list of tuples containing payment data.
        """
        wb = self.load_workbook()
        ws = wb.active
        payments = []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            payments.append(row)

        wb.close()
        return payments    

    def get_payment_counts(self):
        """
        Counts the number of Paid and Pending payments.
        Returns (total_paid, total_pending).
        """
        wb = self.load_workbook()
        ws = wb.active

        total_paid = 0
        total_pending = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[-1] == "Paid":
                total_paid += 1
            elif row[-1] == "Pending":
                total_pending += 1

        wb.close()
        return total_paid, total_pending  

    def generate_payment_chart(self):
        """
        Generates a pie chart showing the ratio of 'Paid' vs 'Pending' payments
        and adds it to the Excel file.
        """
        wb = self.load_workbook()
        ws = wb.active

        # Count the number of 'Paid' and 'Pending' payments
        total_paid = 0
        total_pending = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[-1] == "Paid":
                total_paid += 1
            elif row[-1] == "Pending":
                total_pending += 1

        # If there are no payments, do not create a chart
        if total_paid == 0 and total_pending == 0:
            print("⚠️ No payments found. Chart will not be created.")
            wb.close()
            return

        # Add data for the chart to a new sheet
        chart_sheet = wb.create_sheet(title="Payment Chart")
        chart_sheet.append(["Status", "Count"])
        chart_sheet.append(["Paid", total_paid])
        chart_sheet.append(["Pending", total_pending])

        # Define chart data range
        data = Reference(chart_sheet, min_col=2, min_row=2, max_row=3)
        labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)

        # Create pie chart
        pie_chart = PieChart()
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(labels)

        pie_chart.title = "Paid vs Pending Payments"
        pie_chart.style = 2
        

        # Optimize chart layout
        pie_chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.1, 
                y=0.07,
                w=0.8, 
                h=0.8 
            )
        )

        # Add chart to sheet
        chart_sheet.add_chart(pie_chart, "E5")

        # Save the workbook
        wb.save(self.file_path)
        wb.close()
        print("✅ Payment chart added to Excel!")

       